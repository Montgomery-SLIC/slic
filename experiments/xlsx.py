import io
import re
import openpyxl
from openpyxl.styles import Font

from responses.models import Response, ClickResponse


def _sheet_name(count, name):
    safe = re.sub(r'[\\/*?:\[\]]', '', name.lower().replace(' ', '_'))
    safe = re.sub(r'[^a-z0-9_]', '', safe)
    raw = f"{count}_{safe}"
    return raw[:31]


def generate_xlsx(experiment):
    """Generate XLSX results workbook for an experiment. Returns BytesIO."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    count = [0]

    def bold_row(ws, row_num=1):
        for cell in ws[row_num]:
            cell.font = Font(bold=True)

    def col_key(prompt):
        return '_' + re.sub(r'[^a-z0-9]', '_', (prompt or '').lower())[:40]

    def question_page(qt, name_override=None):
        count[0] += 1
        ws_name = _sheet_name(count[0], name_override or qt.task_ptr.name)
        ws = wb.create_sheet(title=ws_name)
        questions = list(qt.questions.order_by('sort'))
        headers = ['_uid'] + [col_key(q.prompt) for q in questions]
        ws.append(headers)
        bold_row(ws)
        pids = (
            Response.objects
            .filter(question__question_task=qt)
            .values_list('participant_id', flat=True)
            .distinct()
            .order_by('participant_id')
        )
        for pid in pids:
            row = [pid]
            for q in questions:
                resp = Response.objects.filter(question=q, participant_id=pid).first()
                row.append(resp.answer if resp else '')
            ws.append(row)

    OLD_ACCIDENT = {'I clicked by accident', 'I clicked at that time by an accident'}
    OLD_DONTKNOW = {"I don't know", "I don&#39;t know", "I don&amp;#39;t know"}

    def parse_click(cr):
        t, pid = cr.time, cr.participant_id
        ans = (cr.answer or '').strip()
        if cr.no_clicks_explanation:
            return [None, pid, ans, None, None, 'true']
        if cr.from_checkbox and ans == 'accident':
            return [t, pid, None, None, 'true', None]
        if cr.from_checkbox and ans == 'dontknow':
            return [t, pid, None, 'true', None, None]
        if ans in OLD_ACCIDENT:
            return [t, pid, None, None, 'true', None]
        if ans in OLD_DONTKNOW:
            return [t, pid, None, 'true', None, None]
        return [t, pid, ans, None, None, None]

    def click_sheet(ct, name_override=None):
        count[0] += 1
        ws_name = _sheet_name(count[0], name_override or ct.task_ptr.name)
        ws = wb.create_sheet(title=ws_name)
        ws.append(['_time', '_uid', '_comment', '_dk', '_accident', '_no_click'])
        bold_row(ws)
        for cr in ClickResponse.objects.filter(click_task=ct).order_by('participant_id', 'time'):
            ws.append(parse_click(cr))

    def sample_page(st):
        for subtask in st.subtasks.order_by('sort'):
            specific = subtask.get_specific()
            if specific is None:
                continue
            combined = f"{subtask.name} {st.task_ptr.name}"
            if hasattr(specific, 'questions'):
                question_page(specific, combined)
            elif hasattr(specific, 'explanation_prompt'):
                click_sheet(specific, combined)

    for task in experiment.tasks.order_by('sort'):
        specific = task.get_specific()
        if specific is None:
            continue
        if hasattr(specific, 'questions'):
            question_page(specific)
        elif hasattr(specific, 'subtasks'):
            sample_page(specific)

    if not wb.sheetnames:
        wb.create_sheet(title='no_data')
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
